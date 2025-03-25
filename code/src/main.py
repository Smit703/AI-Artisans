from flask import Flask, request, render_template, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import google.generativeai as genai
import os
import pandas as pd

script_dir = os.path.dirname(os.path.abspath(__file__))
Test_case_file = os.path.join(script_dir, "Generated/test_cases.txt")
excel_file_path = os.path.join(script_dir, "Generated/test_cases.xlsx")

app = Flask(__name__)

def load_config():
    file_path = os.path.join(script_dir, "config.json")
    with open(file_path, "r") as config_file:
        return json.load(config_file)

def load_instructions():
    file_path = os.path.join(script_dir, "instructions.md")
    with open(file_path, "r") as instructions_file:
        return instructions_file.read()

def generate_tests_prompt(context, num_cases=10):
    instructions = load_instructions()
    return f"""
    You are an expert in financial transactions and risk assessment. Generate {num_cases} test cases for the following scenario:
    
    Scenario: {context}
    
    Based on the context, follow these instructions:
    {instructions}
    
    Each test case should include:
    - Test Case ID
    - Test scenario (gherkin format, dont use '<br>' for line breaks)
    - Test data (use '<br>' for line breaks)
    - Validation steps (use '<br>' for line breaks)
    - Expected Results (use '<br>' for line breaks)
    
    Validate the test cases for accuracy, completeness, and relevance.
    
    Analyze and add:
    - Missing test scenarios and edge cases
    
    Format the output as a table with the following columns:
    | Test Case ID | Test Scenario | Test Data | Validation Steps | Expected Results
    """

def generate_test_cases(context, num_cases=5):
    prompt = generate_tests_prompt(context, num_cases)
    CONFIG = load_config()
    API_KEY = CONFIG["gemini_api_key"]
    genai.configure(api_key=API_KEY)
    model = genai.GenerativeModel("gemini-2.0-flash")
    
    try:
        response = model.generate_content(prompt)
        test_cases = response.text.strip()
        save_to_excel(test_cases, context)
        return jsonify({"message": "Test cases generated successfully."})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)})

def save_to_txt(test_cases):
    # Ensure the test cases are saved in a readable table format
    with open(Test_case_file, "w") as file:
        file.write("Test Cases in Table Format:\n")
        file.write(test_cases)

def save_to_excel(test_cases, context):
    # Parse the test cases into a DataFrame and save them as an Excel file
    rows = []
    lines = test_cases.split("\n")
    headers = ["Test Case ID", "Test Scenario", "Test Data", "Validation Steps", "Expected Results"]
    
    # Skip the header row and parse the table rows
    for line in lines[4:]:
        if line.strip() and line.count("|") >= len(headers) + 1:  # Ensure the line has enough columns
            columns = [cell.strip() for cell in line.split("|")[1:-1]]  # Remove leading/trailing pipes
           
            # Bolden Given, When, Then in Test Scenario
            columns[1] = columns[1].replace("Given", "**Given**").replace("When", "**When**").replace("Then", "**Then**")
           
            # Add numbering to Test Data, Validation Steps, and Expected Output
            for i in range(2, 5):  # Columns 2, 3, and 4
                if "<br>" in columns[i]:  # Only process if there are multiple lines
                    lines_in_column = columns[i].split("<br>")
                    # Check if the line is already numbered to avoid duplication
                    numbered_lines = [
                        f"{idx + 1}. {line.strip()}" if not line.strip().startswith(f"{idx + 1}.") else line.strip()
                        for idx, line in enumerate(lines_in_column) if line.strip()
                    ]
                    columns[i] = "\n".join(numbered_lines)

            rows.append(columns)

    # Create a DataFrame and save it as an Excel file
    df = pd.DataFrame(rows, columns=headers)
    df.to_excel(excel_file_path, index=False, engine='openpyxl')
    
    # Open the Excel file and apply text wrapping to all cells
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Add a heading at the top
    heading = f"Context based Test Scenarios for the given requirement <{context}>"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    heading_cell = sheet.cell(row=1, column=1)
    heading_cell.value = heading
    heading_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold fill
    heading_cell.font = Font(size=14, bold=True)
    heading_cell.alignment = Alignment(horizontal="center", vertical="center")
    
   # Write headers explicitly in the 2nd row
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(bold=True)

    # Define the table range to include all rows and columns with data
    table_range = f"A1:{chr(64 + len(headers))}{len(rows) + 1}"  # Adjust range based on the number of rows and columns
    table = Table(displayName="TestCasesTable", ref=table_range)

    # Apply a table style
    style = TableStyleInfo(
        name="TableStyleMedium9",  # Predefined style with light blue headers
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Apply text wrapping to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

   # Apply black borders to all cells in the table
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for row in sheet.iter_rows(min_row=1, max_row=len(rows) + 2, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Save the updated Excel file
    workbook.save(excel_file_path)
    print(f"Test cases saved to {excel_file_path}")        

@app.route("/", methods=["GET", "POST"])
def index():
    message = None # Retrieve and clear the message 
    if request.method == "POST":
        context = request.form.get("context")
        num_cases = int(request.form.get("num_cases", 10))
        try:
            generate_test_cases(context, num_cases)  # Store the message in the session
        except Exception as e:
            message = f"Error: {str(e)}"  # Store the error message in the session
    return render_template("index.html", message=message)

@app.route("/download")
def download_file():
    return send_file(excel_file_path, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=False)